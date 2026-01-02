"""
FOB Analysis Script
===================
Automates the processing of RJO Daily FOB Comparative PDF reports.

Process:
1. Converts daily PDFs to individual Excel files in the csv/ directory.
2. Updates a master summary.xlsx workbook with new data from the daily Excel files.

Author: Created for Chris Zitlow FOB Data Processing
Date: October 16, 2025
"""


# These are my imports
import os
import re
import pandas as pd
import pdfplumber
import openpyxl
from openpyxl.utils import range_boundaries
from datetime import datetime

# Configuration
DATA_DIR = 'data'
CSV_DIR = 'csv'
SUMMARY_FILE = 'summary.xlsx'

# Ordered list of sheets/products as they appear in the PDF (Left->Right, Top->Bottom)
SHEET_NAMES = [
    # Row 1
    "SRW_US_Gulf", "HRW_Texas_Gulf", "HRW_US_PNW", "White_Wheat_US_PNW",
    # Row 2
    "DNS_US_PNW", "No1_Milling_DNS", "Argentine_Wheat", "French_Wheat",
    # Row 3
    "Russian_Milling_Wheat", "Ukrainian_Milling_Wheat", "Russian_Corn", "Ukrainian_Corn",
    # Row 4
    "Gulf_Corn", "PNW_Corn", "Argentine_Corn", "Brazilian_Corn",
    # Row 5
    "Gulf_Soybeans", "Argentine_Soybeans", "Brazilian_Soybeans", "Canadian_Canola",
    # Row 6
    "Gulf_Soyoil", "Argentine_Soyoil", "Brazilian_Soyoil", "Malaysian_Olein",
    # Row 7
    "Gulf_Soymeal", "Argentine_Soymeal_Pellets", "Brazilian_Soymeal_Pellets", "Indian_Soymeal"
]

def ensure_directories():
    """Ensure data and csv directories exist."""
    if not os.path.exists(DATA_DIR):
        os.makedirs(DATA_DIR)
    if not os.path.exists(CSV_DIR):
        os.makedirs(CSV_DIR)

def get_date_from_filename(filename):
    """Extract date from filename like RJODailyFOBComparative2025-10-15202510.pdf"""
    match = re.search(r'RJODailyFOBComparative(\d{4}-\d{2}-\d{2})', filename)
    if match:
        return match.group(1)
    return None

def parse_pdf_to_dfs(pdf_path, date_str):
    """
    Parses the PDF and returns a dictionary of {sheet_name: dataframe}.
    """
    print(f"Parsing PDF: {pdf_path}")
    
    with pdfplumber.open(pdf_path) as pdf:
        if not pdf.pages:
            print("  No pages found in PDF.")
            return {}
        
        page = pdf.pages[0]
        tables = page.extract_tables()
        
        if not tables:
            print("  No tables found on first page.")
            return {}
        
        raw_data = tables[0]

    # Identify the 7 sections
    # We look for rows where the second column (index 1) has text, which indicates a header row.
    # The structure is: Header Row, Subheader Row, Data Rows...
    
    section_starts = []
    for i, row in enumerate(raw_data):
        # Heuristic: Header rows have text in the second column (index 1) 
        # and usually empty first column (index 0)
        if i < len(raw_data) and row[1] and str(row[1]).strip():
            # Check if it's a subheader row (contains 'vs')
            if 'vs' in str(row[1]):
                continue
            # Check if it's a data row (Month name in first col)
            if row[0] and str(row[0]).strip():
                continue
                
            section_starts.append(i)

    if len(section_starts) != 7:
        print(f"  Warning: Found {len(section_starts)} sections, expected 7. Trying to proceed...")

    results = {}
    
    # Process each section
    for section_idx, start_row in enumerate(section_starts):
        # Determine end of this section
        if section_idx < len(section_starts) - 1:
            end_row = section_starts[section_idx + 1]
        else:
            end_row = len(raw_data)
            
        section_data = raw_data[start_row:end_row]
        
        # Row 0: Headers (Product Names) - We use the predefined SHEET_NAMES instead
        # Row 1: Subheaders (vs, Chg, Flat, Chg)
        # Row 2+: Data
        
        # There are 4 products per section
        # Columns: 0=Month, 1-4=Prod1, 5-8=Prod2, 9-12=Prod3, 13-16=Prod4
        
        data_rows = section_data[2:]
        
        # Filter valid data rows (must have a month in col 0)
        # Also filter out disclaimer/footer rows which might appear in the last section
        valid_rows = []
        for r in data_rows:
            if not r[0] or not str(r[0]).strip():
                continue
            
            val = str(r[0]).strip()
            # Filter out disclaimer or footer text
            # Month names are short. Disclaimer text is long or contains specific keywords.
            if "Price" in val or "Disclaimer" in val or len(val) > 15:
                continue
                
            valid_rows.append(r)
        
        for i in range(4):
            # Calculate global product index
            prod_idx = (section_idx * 4) + i
            if prod_idx >= len(SHEET_NAMES):
                break
                
            sheet_name = SHEET_NAMES[prod_idx]
            
            # Extract columns for this product
            # Col 0 is Month
            # Cols for product: 1 + (i*4) to 1 + (i*4) + 4
            col_start = 1 + (i * 4)
            col_end = col_start + 4
            
            # Build list of dicts
            clean_data = []
            for row in valid_rows:
                if len(row) > col_start:
                    # Get the chunk
                    chunk = row[col_start:col_end]
                    # Pad if necessary
                    while len(chunk) < 4:
                        chunk.append(None)
                        
                    entry = {
                        'Month': str(row[0]).strip(),
                        'vs': chunk[0],
                        'Chg1': chunk[1],
                        'Flat': chunk[2],
                        'Chg2': chunk[3],
                        'Date': datetime.strptime(date_str, '%Y-%m-%d')
                    }
                    clean_data.append(entry)
            
            df = pd.DataFrame(clean_data)
            # Rename 'vs' column to include the comparison target if possible, 
            # but user requested specific format: Month, vs W, Chg1, Flat, Chg2, Date
            # The subheader row (row[1]) has the specific 'vs W', 'vs KW' etc.
            # Let's try to get the specific 'vs' header from the subheader row
            
            vs_header = 'vs'
            if len(section_data) > 1 and len(section_data[1]) > col_start:
                val = section_data[1][col_start]
                if val:
                    vs_header = str(val).strip()
            
            if not df.empty:
                df = df.rename(columns={'vs': vs_header})
                
            results[sheet_name] = df

    return results

def process_daily_pdfs():
    """Step 1: Iterate PDFs and create Excel files."""
    print("\n--- Step 1: Processing Daily PDFs ---")
    
    files = [f for f in os.listdir(DATA_DIR) if f.endswith('.pdf')]
    
    for pdf_file in files:
        date_str = get_date_from_filename(pdf_file)
        if not date_str:
            print(f"Skipping {pdf_file}: Could not parse date.")
            continue
            
        # Format output filename: FOB_%Y%M%D.xlsx -> FOB_YYYYMMDD.xlsx
        # User requested FOB_%Y%M%D, assuming YYYYMMDD
        date_obj = datetime.strptime(date_str, '%Y-%m-%d')
        output_filename = f"FOB_{date_obj.strftime('%Y%m%d')}.xlsx"
        output_path = os.path.join(CSV_DIR, output_filename)
        
        if os.path.exists(output_path):
            # print(f"Skipping {pdf_file}: {output_filename} already exists.")
            continue
            
        print(f"Processing {pdf_file} -> {output_filename}")
        
        try:
            pdf_path = os.path.join(DATA_DIR, pdf_file)
            dfs = parse_pdf_to_dfs(pdf_path, date_str)
            
            if dfs:
                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    for sheet_name, df in dfs.items():
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                print(f"  Created {output_filename}")
            else:
                print("  Failed to extract data.")
                
        except Exception as e:
            print(f"  Error processing {pdf_file}: {e}")

def update_summary_workbook():
    """Steps 2-5: Update summary.xlsx with missing data."""
    print("\n--- Step 2-5: Updating Summary Workbook ---")
    
    if not os.path.exists(SUMMARY_FILE):
        print(f"Error: {SUMMARY_FILE} not found.")
        return

    # Load workbook
    try:
        wb = openpyxl.load_workbook(SUMMARY_FILE)
    except Exception as e:
        print(f"Error loading {SUMMARY_FILE}: {e}")
        return

    # Get list of available daily Excel files and their dates
    daily_files = {} # date_str -> filepath
    for f in os.listdir(CSV_DIR):
        if f.startswith('FOB_') and f.endswith('.xlsx'):
            # Parse date from FOB_YYYYMMDD.xlsx
            match = re.search(r'FOB_(\d{8})', f)
            if match:
                d_str = match.group(1)
                # Convert to datetime object for comparison
                d_obj = datetime.strptime(d_str, '%Y%m%d')
                daily_files[d_obj] = os.path.join(CSV_DIR, f)

    # Iterate through each sheet in our list
    for sheet_name in SHEET_NAMES:
        if sheet_name not in wb.sheetnames:
            print(f"Sheet '{sheet_name}' not found in summary workbook. Skipping.")
            continue
            
        ws = wb[sheet_name]
        
        # Find the table
        if not ws.tables:
            print(f"No table found in sheet '{sheet_name}'. Skipping.")
            continue
            
        # Assume table name matches sheet name or is the only table
        # User said: "Each sheet... has a table that has the same name as the sheet"
        table = None
        if sheet_name in ws.tables:
            table = ws.tables[sheet_name]
        else:
            # Fallback: take first table
            table = list(ws.tables.values())[0]
            
        # Read table data to DataFrame
        # Get range
        ref = table.ref
        min_col, min_row, max_col, max_row = range_boundaries(ref)
        
        data = ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True)
        data = list(data)
        
        if not data:
            print(f"Empty table in {sheet_name}.")
            continue
            
        headers = data[0]
        rows = data[1:]
        
        df = pd.DataFrame(rows, columns=headers)
        
        # Check for 'Date' column
        if 'Date' not in df.columns:
            print(f"No 'Date' column in {sheet_name}. Skipping.")
            continue
            
        # Ensure Date column is datetime
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce')

        # Get existing dates
        existing_dates = set(df['Date'].dropna().tolist())
        
        # Find missing dates
        updates_needed = []
        for d_obj, file_path in daily_files.items():
            if d_obj not in existing_dates:
                updates_needed.append((d_obj, file_path))
        
        if not updates_needed:
            # print(f"No updates needed for {sheet_name}.")
            continue
            
        print(f"Updating {sheet_name} with {len(updates_needed)} new dates...")
        
        # Load new data
        new_rows = []
        for d_obj, file_path in updates_needed:
            try:
                daily_df = pd.read_excel(file_path, sheet_name=sheet_name)
                
                # Ensure Date is datetime in new data
                if 'Date' in daily_df.columns:
                    daily_df['Date'] = pd.to_datetime(daily_df['Date'], errors='coerce')

                # Ensure columns match
                # The daily df might have different column names for 'vs ...'
                # We need to align them.
                # The summary table headers are fixed.
                # The daily file headers are: Month, vs X, Chg1, Flat, Chg2, Date
                # We assume the structure is compatible (6 columns)
                
                if len(daily_df.columns) == len(df.columns):
                    # Rename columns to match summary df to allow concat
                    daily_df.columns = df.columns
                    new_rows.append(daily_df)
                else:
                    print(f"  Column mismatch in {file_path} for {sheet_name}")
            except Exception as e:
                print(f"  Error reading {file_path}: {e}")
        
        if new_rows:
            # Append
            new_data = pd.concat(new_rows, ignore_index=True)
            df = pd.concat([df, new_data], ignore_index=True)
            
            # Remove empty rows
            df = df.dropna(how='all')
            if 'Month' in df.columns:
                df = df[df['Month'].notna() & (df['Month'].astype(str).str.strip() != '')]

            # Sort by Date
            df = df.sort_values('Date')
            
            # Write back to Excel
            # Clear existing data rows (keep header)
            # We need to be careful not to destroy the table structure
            # openpyxl doesn't easily support "clearing table data" while keeping style
            # Strategy: Write over the cells, then update table range
            
            # Convert df to rows
            updated_data = [list(df.columns)] + df.values.tolist()
            
            # Write data
            for r_idx, row_data in enumerate(updated_data):
                for c_idx, value in enumerate(row_data):
                    ws.cell(row=min_row + r_idx, column=min_col + c_idx, value=value)
            
            # Update table reference
            new_max_row = min_row + len(updated_data) - 1
            new_ref = f"{openpyxl.utils.get_column_letter(min_col)}{min_row}:{openpyxl.utils.get_column_letter(max_col)}{new_max_row}"
            table.ref = new_ref
            
            print(f"  Updated {sheet_name}. New row count: {len(df)}")

    # Save workbook
    print("Saving summary workbook...")
    wb.save(SUMMARY_FILE)
    print("Done.")

def main():
    ensure_directories()
    process_daily_pdfs()
    update_summary_workbook()

if __name__ == "__main__":
    main()
